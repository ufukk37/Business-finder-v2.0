"""
Dışa Aktarım API Endpoint'leri
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
import io
import json
from datetime import datetime

from app.core.database import get_db
from app.models.business import Business

router = APIRouter()


@router.get("/download/{format}")
def download_export(
    format: str,
    ids: Optional[str] = Query(None, description="Virgülle ayrılmış ID'ler"),
    city: Optional[str] = None,
    business_type: Optional[str] = None,
    has_phone: Optional[bool] = None,
    has_website: Optional[bool] = None,
    min_rating: Optional[float] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Filtrelenmiş işletmeleri dışa aktar"""
    
    if format not in ["xlsx", "csv", "json"]:
        raise HTTPException(status_code=400, detail="Geçersiz format. xlsx, csv veya json kullanın.")
    
    # Query oluştur
    query = db.query(Business)
    
    # ID bazlı filtreleme
    if ids:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip().isdigit()]
        if id_list:
            query = query.filter(Business.id.in_(id_list))
    else:
        # Diğer filtreler
        if city:
            query = query.filter(Business.city.ilike(f"%{city}%"))
        if business_type:
            query = query.filter(Business.business_type.ilike(f"%{business_type}%"))
        if has_phone:
            query = query.filter(Business.phone.isnot(None), Business.phone != "")
        if has_website:
            query = query.filter(Business.website.isnot(None), Business.website != "")
        if min_rating:
            query = query.filter(Business.rating >= min_rating)
        if search:
            query = query.filter(
                (Business.name.ilike(f"%{search}%")) |
                (Business.address.ilike(f"%{search}%"))
            )
    
    businesses = query.all()
    
    if not businesses:
        raise HTTPException(status_code=404, detail="İşletme bulunamadı")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if format == "xlsx":
        return _generate_xlsx(businesses, timestamp)
    elif format == "csv":
        return _generate_csv(businesses, timestamp)
    else:
        return _generate_json(businesses, timestamp)


def _generate_xlsx(businesses: List[Business], timestamp: str):
    """Excel dosyası oluştur"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl kütüphanesi yüklü değil")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "İşletmeler"
    
    # Başlık stilleri
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = [
        "ID", "İşletme Adı", "Adres", "Şehir", "İlçe", 
        "Telefon", "Website", "Kategori", "Puan", 
        "Değerlendirme", "Notlar", "Etiketler"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Veriler
    for row, business in enumerate(businesses, 2):
        data = [
            business.id,
            business.name,
            business.address or "",
            business.city or "",
            business.district or "",
            business.phone or "",
            business.website or "",
            business.business_type or "",
            business.rating or "",
            business.total_ratings or 0,
            business.notes or "",
            business.tags or ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Sütun genişlikleri
    column_widths = [8, 35, 45, 15, 15, 18, 35, 15, 8, 12, 30, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Dosyayı oluştur
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=isletmeler_{timestamp}.xlsx"
        }
    )


def _generate_csv(businesses: List[Business], timestamp: str):
    """CSV dosyası oluştur"""
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    
    # Başlıklar
    writer.writerow([
        "ID", "İşletme Adı", "Adres", "Şehir", "İlçe",
        "Telefon", "Website", "Kategori", "Puan",
        "Değerlendirme", "Notlar", "Etiketler"
    ])
    
    # Veriler
    for business in businesses:
        writer.writerow([
            business.id,
            business.name,
            business.address or "",
            business.city or "",
            business.district or "",
            business.phone or "",
            business.website or "",
            business.business_type or "",
            business.rating or "",
            business.total_ratings or 0,
            business.notes or "",
            business.tags or ""
        ])
    
    output.seek(0)
    content = output.getvalue().encode('utf-8-sig')
    
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=isletmeler_{timestamp}.csv"
        }
    )


def _generate_json(businesses: List[Business], timestamp: str):
    """JSON dosyası oluştur"""
    
    data = [business.to_dict() for business in businesses]
    
    content = json.dumps(data, ensure_ascii=False, indent=2)
    
    return StreamingResponse(
        io.BytesIO(content.encode('utf-8')),
        media_type="application/json; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=isletmeler_{timestamp}.json"
        }
    )
